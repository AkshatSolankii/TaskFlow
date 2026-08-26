from flask import Blueprint, request, jsonify, send_file, Response
from sqlalchemy import case, func
from datetime import datetime
from flask_login import login_required, current_user
from models import db, Task, Category, ActivityLog, TaskMember, can_access_task
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

task_bp = Blueprint('tasks', __name__)

EXPORT_HEADERS = ["Task", "Description", "Due Date", "Priority", "Status", "Category"]


def _build_export_query():
    status        = request.args.get('status',        '', type=str).strip()
    priority      = request.args.get('priority',      '', type=str).strip()
    category_id   = request.args.get('category_id',  None, type=int)
    category_name = request.args.get('category_name', '', type=str).strip()
    sort          = request.args.get('sort',          '', type=str).strip()
    search        = request.args.get('search',        '', type=str).strip()

    # RBAC: Admin/Manager export across ALL users' tasks; User exports only their own.
    if current_user.can_manage_all_tasks():
        query = Task.query
    else:
        query = Task.query.filter_by(user_id=current_user.id)

    if search:
        query = query.filter(
            Task.title.ilike(f'%{search}%') |
            Task.description.ilike(f'%{search}%')
        )
    if status:
        query = query.filter(Task.status == status)
    if priority:
        query = query.filter(Task.priority == priority)
    if category_id:
        query = query.filter(Task.category_id == category_id)
    elif category_name:
        # Categories are global now, so no per-user scoping when looking one up by name.
        category = Category.query.filter(Category.name.ilike(category_name)).first()
        cat_id = category.id if category else -1
        query = query.filter(Task.category_id == cat_id)

    if sort == "priority":
        priority_order = case(
            (Task.priority == "High",   1),
            (Task.priority == "Medium", 2),
            (Task.priority == "Low",    3),
            else_=4
        )
        query = query.order_by(priority_order)
    else:
        query = query.order_by(Task.created_at.desc())

    return query


def _task_row(task):
    return [
        task.title,
        task.description or "",
        task.deadline    or "",
        task.priority,
        task.status,
        task.category.name if task.category else ""
    ]


def _can_modify_task(task):
    """RBAC + task-level check for write operations (update/delete)."""
    if current_user.can_manage_all_tasks():
        return True
    return task.user_id == current_user.id


# ================= CREATE TASK =================
@task_bp.route('/tasks', methods=['POST'])
@login_required
def create_task():

    data = request.get_json()

    if not data or 'title' not in data:
        return jsonify({"error": "Title is required"}), 400

    new_task = Task(
        title       = data['title'],
        description = data.get('description', ''),
        deadline    = data.get('deadline'),
        priority    = data.get('priority', 'Medium'),
        status      = data.get('status', 'pending'),
        user_id     = current_user.id,
        category_id = data.get('category_id')
    )

    db.session.add(new_task)
    db.session.commit()

    activity = ActivityLog(
        action      = "Created",
        entity_type = "Task",
        entity_name = new_task.title,
        user_id     = current_user.id
    )
    db.session.add(activity)
    db.session.commit()

    return jsonify({
        "message": "Task created successfully",
        "task":    new_task.to_dict()
    }), 201


# ================= GET TASKS =================
@task_bp.route('/tasks', methods=['GET'])
@login_required
def get_tasks():

    page  = request.args.get('page',  1,  type=int)
    limit = request.args.get('limit', 10, type=int)
    sort  = request.args.get('sort',  '', type=str)

    # RBAC: Admin/Manager see ALL tasks in "tasks"; User sees only their own,
    # same as before RBAC existed.
    if current_user.can_manage_all_tasks():
        query = Task.query
    else:
        query = Task.query.filter_by(user_id=current_user.id)

    if sort == "priority":
        priority_order = case(
            (Task.priority == "High",   1),
            (Task.priority == "Medium", 2),
            (Task.priority == "Low",    3),
            else_=4
        )
        query = query.order_by(priority_order)
    else:
        query = query.order_by(Task.created_at.desc())

    pagination = query.paginate(page=page, per_page=limit, error_out=False)

    # ── Shared tasks (accepted invitations) — unchanged, User-scoped ──
    # Admin/Manager already see everything via the query above, so we
    # don't need to also compute shared_tasks for them (would duplicate).
    shared_tasks = []
    if not current_user.can_manage_all_tasks():
        memberships = TaskMember.query.filter_by(user_id=current_user.id).all()
        task_ids    = [m.task_id for m in memberships]

        if task_ids:
            tasks_by_id = {
                t.id: t for t in Task.query.filter(Task.id.in_(task_ids)).all()
            }
            for m in memberships:
                task = tasks_by_id.get(m.task_id)
                if task:
                    d = task.to_dict()
                    d["shared_role"]    = m.role
                    d["owner_username"] = task.user.username
                    d["is_shared"]      = True
                    shared_tasks.append(d)

    return jsonify({
        "tasks":        [t.to_dict() for t in pagination.items],
        "shared_tasks": shared_tasks,
        "page":         pagination.page,
        "total_pages":  pagination.pages,
        "total_tasks":  pagination.total
    }), 200


# ================= CALENDAR TASKS =================
@task_bp.route('/calendar-tasks', methods=['GET'])
@login_required
def get_calendar_tasks():
    """Return every due-dated task the current user is allowed to view."""
    if current_user.can_manage_all_tasks():
        tasks = Task.query.filter(Task.deadline.isnot(None)).all()
    else:
        shared_ids = [
            membership.task_id
            for membership in TaskMember.query.filter_by(user_id=current_user.id).all()
        ]
        query = Task.query.filter(Task.deadline.isnot(None))
        if shared_ids:
            query = query.filter(
                (Task.user_id == current_user.id) | Task.id.in_(shared_ids)
            )
        else:
            query = query.filter_by(user_id=current_user.id)
        tasks = query.all()

    return jsonify({"tasks": [task.to_dict() for task in tasks]}), 200


# ================= GET SINGLE TASK =================
@task_bp.route('/tasks/<int:id>', methods=['GET'])
@login_required
def get_task(id):

    task = Task.query.get(id)
    if not task:
        return jsonify({"error": "Task not found"}), 404

    # RBAC: Admin/Manager can view any task, bypassing per-task TaskMember checks
    if current_user.can_manage_all_tasks():
        data = task.to_dict()
        data["your_role"]      = current_user.role  # "Admin" | "Manager"
        data["owner_username"] = task.user.username
        return jsonify(data), 200

    allowed, role = can_access_task(current_user, task)
    if not allowed:
        return jsonify({"error": "Forbidden"}), 403

    data = task.to_dict()
    data["your_role"]      = role
    data["owner_username"] = task.user.username
    return jsonify(data), 200


# ================= EXPORT TASKS (CSV) =================
@task_bp.route('/tasks/export/csv', methods=['GET'])
@login_required
def export_tasks_csv():

    tasks = _build_export_query().all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(EXPORT_HEADERS)
    for task in tasks:
        writer.writerow(_task_row(task))

    csv_bytes = output.getvalue().encode('utf-8')

    activity = ActivityLog(action="Exported", entity_type="Task",
                           entity_name="CSV Export", user_id=current_user.id)
    db.session.add(activity)
    db.session.commit()

    return Response(csv_bytes, status=200, headers={
        "Content-Disposition": "attachment; filename=tasks.csv",
        "Content-Type":        "text/csv; charset=utf-8"
    })


# ================= EXPORT TASKS (EXCEL) =================
@task_bp.route('/tasks/export/xlsx', methods=['GET'])
@task_bp.route('/tasks/export/excel', methods=['GET'])
@login_required
def export_tasks_excel():

    tasks = _build_export_query().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="4F81BD")
    header_align = Alignment(horizontal="center", vertical="center")

    ws.append(EXPORT_HEADERS)
    for col_idx, _ in enumerate(EXPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    for task in tasks:
        ws.append(_task_row(task))

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    file = io.BytesIO()
    wb.save(file)
    file.seek(0)

    activity = ActivityLog(action="Exported", entity_type="Task",
                           entity_name="Excel Export", user_id=current_user.id)
    db.session.add(activity)
    db.session.commit()

    return send_file(file, as_attachment=True, download_name="tasks.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ================= UPDATE TASK =================
@task_bp.route('/tasks/<int:id>', methods=['PATCH'])
@login_required
def update_task(id):

    task = Task.query.get(id)
    if not task:
        return jsonify({"error": "Task not found"}), 404

    # RBAC: Admin/Manager can update any task.
    if current_user.can_manage_all_tasks():
        is_owner  = True   # treat as owner-equivalent for the category-edit check below
        is_editor = True
    else:
        is_owner  = task.user_id == current_user.id
        member    = TaskMember.query.filter_by(
            task_id=id, user_id=current_user.id
        ).first()
        is_editor = member and member.role == "Editor"

        if not (is_owner or is_editor):
            return jsonify({"error": "Forbidden — Editors and Owners only"}), 403

    data = request.get_json()

    task.title       = data.get("title",       task.title)
    task.description = data.get("description", task.description)
    task.deadline    = data.get("deadline",    task.deadline)
    task.priority    = data.get("priority",    task.priority)
    task.status      = data.get("status",      task.status)

    if is_owner:
        task.category_id = data.get("category_id", task.category_id)

    activity = ActivityLog(
        action      = "Updated",
        entity_type = "Task",
        entity_name = task.title,
        user_id     = current_user.id
    )
    db.session.add(activity)
    db.session.commit()

    return jsonify({
        "message": "Task updated successfully",
        "task":    task.to_dict()
    }), 200


# ================= DELETE TASK =================
@task_bp.route('/tasks/<int:id>', methods=['DELETE'])
@login_required
def delete_task(id):

    task = Task.query.get(id)
    if not task:
        return jsonify({"error": "Task not found or not authorised"}), 404

    if not _can_modify_task(task):
        return jsonify({"error": "Task not found or not authorised"}), 404

    task_name = task.title

    activity = ActivityLog(
        action      = "Deleted",
        entity_type = "Task",
        entity_name = task_name,
        user_id     = current_user.id
    )
    db.session.add(activity)
    db.session.delete(task)
    db.session.commit()

    return jsonify({"message": "Task deleted successfully"}), 200


# ================= BULK: MARK STATUS =================
@task_bp.route('/tasks/bulk-status', methods=['POST'])
@login_required
def bulk_update_status():
    data   = request.get_json() or {}
    ids    = data.get('ids', [])
    status = (data.get('status') or '').strip()

    if not isinstance(ids, list) or not ids:
        return jsonify({"error": "No task ids provided"}), 400
    if status not in ("pending", "in_progress", "completed"):
        return jsonify({"error": "Invalid status"}), 400

    query = Task.query.filter(Task.id.in_(ids))
    if not current_user.can_manage_all_tasks():
        query = query.filter(Task.user_id == current_user.id)
    tasks = query.all()

    if not tasks:
        return jsonify({"error": "No matching tasks found"}), 404

    for task in tasks:
        task.status = status

    activity = ActivityLog(
        action      = f"Bulk marked {len(tasks)} task(s) as {status}",
        entity_type = "Task",
        entity_name = f"{len(tasks)} tasks",
        user_id     = current_user.id
    )
    db.session.add(activity)
    db.session.commit()

    return jsonify({"message": f"{len(tasks)} task(s) updated", "updated": len(tasks)}), 200


# ================= BULK: CHANGE CATEGORY =================
@task_bp.route('/tasks/bulk-category', methods=['POST'])
@login_required
def bulk_update_category():
    data        = request.get_json() or {}
    ids         = data.get('ids', [])
    category_id = data.get('category_id')

    # Validate task ids FIRST — no point checking the category
    # if the request doesn't even name any tasks to update.
    if not isinstance(ids, list) or not ids:
        return jsonify({"error": "No task ids provided"}), 400

    if category_id:
        # Categories are global now — just confirm it exists.
        category = Category.query.filter_by(id=category_id).first()
        if not category:
            return jsonify({"error": "Category not found"}), 404

    query = Task.query.filter(Task.id.in_(ids))
    if not current_user.can_manage_all_tasks():
        query = query.filter(Task.user_id == current_user.id)
    tasks = query.all()

    if not tasks:
        return jsonify({"error": "No matching tasks found"}), 404

    for task in tasks:
        task.category_id = category_id

    activity = ActivityLog(
        action      = f"Bulk changed category for {len(tasks)} task(s)",
        entity_type = "Task",
        entity_name = f"{len(tasks)} tasks",
        user_id     = current_user.id
    )
    db.session.add(activity)
    db.session.commit()

    return jsonify({"message": f"{len(tasks)} task(s) updated", "updated": len(tasks)}), 200


# ================= BULK: DELETE =================
@task_bp.route('/tasks/bulk-delete', methods=['POST'])
@login_required
def bulk_delete_tasks():
    data = request.get_json() or {}
    ids  = data.get('ids', [])

    if not isinstance(ids, list) or not ids:
        return jsonify({"error": "No task ids provided"}), 400

    query = Task.query.filter(Task.id.in_(ids))
    if not current_user.can_manage_all_tasks():
        query = query.filter(Task.user_id == current_user.id)
    tasks = query.all()

    if not tasks:
        return jsonify({"error": "No matching tasks found"}), 404

    deleted = len(tasks)
    for task in tasks:
        db.session.delete(task)

    activity = ActivityLog(
        action      = f"Bulk deleted {deleted} task(s)",
        entity_type = "Task",
        entity_name = f"{deleted} tasks",
        user_id     = current_user.id
    )
    db.session.add(activity)
    db.session.commit()

    return jsonify({"message": f"{deleted} task(s) deleted", "deleted": deleted}), 200


# ================= CREATE CATEGORY =================
# Categories are now GLOBAL (shared across every user), not per-user.
# Only Admin/Manager may create, rename, or delete a category.
@task_bp.route('/categories', methods=['POST'])
@login_required
def create_category():
    if not current_user.can_manage_all_tasks():
        return jsonify({"error": "Only Admins and Managers can create categories"}), 403

    data = request.get_json()
    name = data.get("name", "").strip()

    if not name:
        return jsonify({"error": "Category name required"}), 400

    existing = Category.query.filter_by(name=name).first()  # global uniqueness now
    if existing:
        return jsonify({"error": "Category already exists"}), 400

    category = Category(name=name, user_id=current_user.id)
    db.session.add(category)
    db.session.commit()

    activity = ActivityLog(action="Created", entity_type="Category",
                           entity_name=category.name, user_id=current_user.id)
    db.session.add(activity)
    db.session.commit()

    return jsonify({"id": category.id, "name": category.name}), 201


# ================= GET CATEGORIES =================
@task_bp.route('/categories', methods=['GET'])
@login_required
def get_categories():
    # Categories are global now — every role can VIEW all categories
    # (to assign them to tasks), even though only Admin/Manager can
    # create/edit/delete them.
    categories = Category.query.all()
    return jsonify([{"id": c.id, "name": c.name} for c in categories])


# ================= UPDATE CATEGORY =================
@task_bp.route('/categories/<int:id>', methods=['PATCH'])
@login_required
def update_category(id):
    if not current_user.can_manage_all_tasks():
        return jsonify({"error": "Only Admins and Managers can edit categories"}), 403

    category = Category.query.get(id)
    if not category:
        return jsonify({"error": "Category not found"}), 404

    data     = request.get_json()
    new_name = data.get("name", "").strip()
    if not new_name:
        return jsonify({"error": "Category name required"}), 400

    category.name = new_name

    activity = ActivityLog(action="Updated", entity_type="Category",
                           entity_name=category.name, user_id=current_user.id)
    db.session.add(activity)
    db.session.commit()

    return jsonify({"message": "Category updated successfully"}), 200


# ================= DELETE CATEGORY =================
@task_bp.route('/categories/<int:id>', methods=['DELETE'])
@login_required
def delete_category(id):
    if not current_user.can_manage_all_tasks():
        return jsonify({"error": "Only Admins and Managers can delete categories"}), 403

    category = Category.query.get(id)
    if not category:
        return jsonify({"error": "Category not found"}), 404

    category_name = category.name

    tasks = Task.query.filter_by(category_id=id).all()  # any task using it, any owner
    for task in tasks:
        task.category_id = None

    activity = ActivityLog(action="Deleted", entity_type="Category",
                           entity_name=category_name, user_id=current_user.id)
    db.session.add(activity)
    db.session.delete(category)
    db.session.commit()

    return jsonify({"message": "Category deleted successfully"}), 200


# ================= ACTIVITY LOG =================
@task_bp.route('/activity-log', methods=['GET'])
@login_required
def activity_log():
    # RBAC: Admin sees everyone's activity; Manager/User see only their own.
    # (Full audit-log visibility is an Admin-tier concern — Task 3's
    # dedicated Audit Logs feature will build on this further.)
    if current_user.is_admin():
        logs = ActivityLog.query.order_by(ActivityLog.timestamp.desc()).all()
    else:
        logs = ActivityLog.query.filter_by(
            user_id=current_user.id
        ).order_by(ActivityLog.timestamp.desc()).all()

    return jsonify([{
        "action":      log.action,
        "entity_type": log.entity_type,
        "entity_name": log.entity_name,
        "user":        log.user.username if current_user.is_admin() else None,
        "timestamp":   log.timestamp.strftime("%Y-%m-%dT%H:%M:%S") + "Z"
    } for log in logs])


# ================= DASHBOARD STATS =================
@task_bp.route('/dashboard-stats', methods=['GET'])
@login_required
def dashboard_stats():

    # RBAC: Admin/Manager see stats across ALL users; User sees only their own.
    base_task_filter = [] if current_user.can_manage_all_tasks() else [Task.user_id == current_user.id]

    status_counts = dict(
        db.session.query(Task.status, func.count(Task.id))
        .filter(*base_task_filter)
        .group_by(Task.status)
        .all()
    )
    total_tasks     = sum(status_counts.values())
    completed_tasks = status_counts.get("completed", 0)
    pending_tasks   = status_counts.get("pending", 0)

    now = datetime.now()
    overdue_tasks = 0
    incomplete_deadlines = (
        db.session.query(Task.deadline)
        .filter(*base_task_filter)
        .filter(Task.status != "completed", Task.deadline.isnot(None))
        .all()
    )
    for (deadline,) in incomplete_deadlines:
        try:
            if datetime.fromisoformat(deadline) < now:
                overdue_tasks += 1
        except Exception:
            pass

    # Categories are global now. For a User (non-manager), we still only want
    # to count *their own* tasks per category, so the join condition itself
    # carries the RBAC filter; for Admin/Manager it's a plain outerjoin.
    if base_task_filter:
        join_condition = (Task.category_id == Category.id) & (Task.user_id == current_user.id)
    else:
        join_condition = Task.category_id == Category.id

    category_rows = (
        db.session.query(Category.id, Category.name, func.count(Task.id))
        .outerjoin(Task, join_condition)
        .group_by(Category.id, Category.name)
        .all()
    )

    category_stats = [
        {"name": name, "count": count} for (_, name, count) in category_rows
    ]

    return jsonify({
        "total_tasks":     total_tasks,
        "completed_tasks": completed_tasks,
        "pending_tasks":   pending_tasks,
        "overdue_tasks":   overdue_tasks,
        "categories":      category_stats
    })


# ================= CURRENT USER =================
@task_bp.route('/me', methods=['GET'])
@login_required
def me():
    return jsonify({
        "id":       current_user.id,
        "username": current_user.username,
        "email":    current_user.email,
        "role":     current_user.role
    }), 200
